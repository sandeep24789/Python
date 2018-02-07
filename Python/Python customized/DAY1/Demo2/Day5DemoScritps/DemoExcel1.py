#import relevant module
import openpyxl

#Opening excel workbook
wb = openpyxl.load_workbook('c:/Training/PyFiles/example.xlsx')

#Accessing a specific sheet
sheet = wb.get_sheet_by_name('Sheet1')

print(sheet.title)

#Different ways to get to cells
print(sheet['A1'])
print(sheet['A1'].value)

print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)
for i in range(1, 8, 2):
     print(i, sheet.cell(row=i, column=2).value)



